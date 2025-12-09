from __future__ import annotations

import csv
import io
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import numpy as np
import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


app = FastAPI(title="Data Insights API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class SchemaGuess(BaseModel):
    date: Optional[str]
    revenue: Optional[str]
    cost: Optional[str]
    profit: Optional[str]
    product: Optional[str]
    qty: Optional[str]


class MetricBreakdown(BaseModel):
    total_revenue: float
    total_cost: float
    total_profit: float
    avg_daily_revenue: float
    weekly_growth_pct: float
    projections: List[float]


class Anomaly(BaseModel):
    date: str
    value: float
    z_score: float
    message: str


class CategoryBreakdown(BaseModel):
    name: str
    total: float


class UploadResponse(BaseModel):
    schema: SchemaGuess
    metrics: MetricBreakdown
    timeseries: List[Dict[str, float]]
    anomalies: List[Anomaly]
    narrative: str
    columns: List[str]
    categories: List[CategoryBreakdown]


class ExplainRequest(BaseModel):
    metrics: MetricBreakdown
    anomalies: List[Anomaly] = []
    categories: List[CategoryBreakdown] = []
    period: Optional[str] = None


class ExplainResponse(BaseModel):
    explanation: str
    tags: List[str]


class PdfRequest(BaseModel):
    metrics: MetricBreakdown
    timeseries: List[Dict[str, float]] = []
    categories: List[CategoryBreakdown] = []
    anomalies: List[Anomaly] = []
    narrative: str


def parse_csv(content: bytes) -> List[Dict[str, str]]:
    text = content.decode("utf-8", errors="ignore")
    try:
        reader = csv.DictReader(io.StringIO(text))
        rows = [row for row in reader if any(row.values())]
    except csv.Error as exc:  # noqa: PERF203
        raise ValueError(
            "Could not parse as CSV. If this is an Excel file, upload .xlsx/.xls directly."
        ) from exc
    if not rows:
        raise ValueError("No rows found in upload")
    return rows


def parse_excel(content: bytes) -> List[Dict[str, str]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Could not read Excel file") from exc
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("No rows found in Excel")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: List[Dict[str, str]] = []
    for raw in rows[1:]:
        record: Dict[str, str] = {}
        for header, value in zip(headers, raw):
            if not header:
                continue
            record[header] = "" if value is None else str(value)
        if any(record.values()):
            data.append(record)
    if not data:
        raise ValueError("No rows found in Excel")
    return data


def apply_overrides(schema: SchemaGuess, overrides: Dict[str, Optional[str]]) -> SchemaGuess:
    return SchemaGuess(
        date=overrides.get("date_column") or schema.date,
        revenue=overrides.get("revenue_column") or schema.revenue,
        cost=schema.cost,
        profit=schema.profit,
        product=overrides.get("category_column") or schema.product,
        qty=schema.qty,
    )


def guess_schema(row: Dict[str, str]) -> SchemaGuess:
    headers = {k.lower(): k for k in row.keys()}
    def find(keys: List[str]) -> Optional[str]:
        for alias in keys:
            for lower, original in headers.items():
                if alias in lower:
                    return original
        return None

    return SchemaGuess(
        date=find(["date", "day"]),
        revenue=find(["revenue", "sales", "gmv"]),
        cost=find(["cost", "cogs", "spend"]),
        profit=find(["profit", "margin"]),
        product=find(["product", "sku", "item", "name"]),
        qty=find(["qty", "quantity", "units"]),
    )


def coerce_float(value: str) -> float:
    try:
        return float(value.replace(",", ""))
    except Exception:
        return 0.0


def coerce_date(value: str) -> Optional[datetime]:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def build_timeseries(rows: List[Dict[str, str]], schema: SchemaGuess) -> List[Dict[str, float]]:
    if not schema.date or not schema.revenue:
        return []

    rollup: Dict[datetime, float] = {}
    for row in rows:
        dt = coerce_date(row.get(schema.date, ""))
        if not dt:
            continue
        revenue = coerce_float(row.get(schema.revenue, "0"))
        rollup[dt] = rollup.get(dt, 0) + revenue

    series = [{"date": dt.strftime("%Y-%m-%d"), "revenue": value} for dt, value in sorted(rollup.items())]
    return series


def build_categories(rows: List[Dict[str, str]], schema: SchemaGuess) -> List[CategoryBreakdown]:
    cat_col = schema.product
    rev_col = schema.revenue
    if not cat_col or not rev_col:
        return []
    agg: Dict[str, float] = {}
    for row in rows:
        key = row.get(cat_col)
        if not key:
            continue
        agg[key] = agg.get(key, 0) + coerce_float(row.get(rev_col, "0"))
    top = sorted(agg.items(), key=lambda item: item[1], reverse=True)[:8]
    return [CategoryBreakdown(name=name, total=round(total, 2)) for name, total in top]


def assemble_payload(rows: List[Dict[str, str]], schema: SchemaGuess) -> UploadResponse:
    metrics = compute_metrics(rows, schema)
    timeseries = build_timeseries(rows, schema)
    anomalies = detect_anomalies(timeseries)
    categories = build_categories(rows, schema)
    narrative = build_narrative(schema, metrics, anomalies)
    columns = list(rows[0].keys())
    return UploadResponse(
        schema=schema,
        metrics=metrics,
        timeseries=timeseries,
        anomalies=anomalies,
        narrative=narrative,
        columns=columns,
        categories=categories,
    )


def detect_anomalies(series: List[Dict[str, float]]) -> List[Anomaly]:
    if len(series) < 5:
        return []
    values = np.array([p["revenue"] for p in series])
    mean = values.mean()
    std = values.std() or 1
    anomalies: List[Anomaly] = []
    for point, raw in zip(series, values):
        z = (raw - mean) / std
        if abs(z) >= 2.5:
            anomalies.append(
                Anomaly(
                    date=point["date"],
                    value=float(raw),
                    z_score=float(round(z, 2)),
                    message=f"Revenue anomaly on {point['date']}: {raw:.2f} (z={z:.2f})",
                )
            )
    return anomalies


def compute_metrics(rows: List[Dict[str, str]], schema: SchemaGuess) -> MetricBreakdown:
    revenue_col = schema.revenue or ""
    profit_col = schema.profit or ""
    cost_col = schema.cost or ""
    date_col = schema.date or ""

    revenues = [coerce_float(r.get(revenue_col, "0")) for r in rows]
    costs = [coerce_float(r.get(cost_col, "0")) for r in rows] if cost_col else [0.0 for _ in rows]
    profits = [coerce_float(r.get(profit_col, "0")) for r in rows] if profit_col else [0.0 for _ in rows]
    dates = [coerce_date(r.get(date_col, "")) for r in rows] if date_col else []

    total_revenue = sum(revenues)
    total_cost = sum(costs)
    total_profit = sum(profits)
    avg_daily_revenue = total_revenue / max(len(set([d.date() for d in dates if d])), 1)

    # Weekly growth: compare last 7 vs prior 7 days where possible
    now = datetime.utcnow().date()
    def window_sum(start_delta: int, end_delta: int) -> float:
        start = now - timedelta(days=start_delta)
        end = now - timedelta(days=end_delta)
        return sum(val for val, dt in zip(revenues, dates) if dt and end < dt.date() <= start)

    recent = window_sum(0, 7)
    previous = window_sum(7, 14)
    weekly_growth_pct = ((recent - previous) / previous * 100) if previous else 0.0

    # Projection: naive mean of last 4 points
    tail = revenues[-4:] if len(revenues) >= 4 else revenues
    mean_tail = (sum(tail) / len(tail)) if tail else 0.0
    projections = [round(mean_tail * (1 + i * 0.02), 2) for i in range(1, 5)]

    return MetricBreakdown(
        total_revenue=round(total_revenue, 2),
        total_cost=round(total_cost, 2),
        total_profit=round(total_profit, 2),
        avg_daily_revenue=round(avg_daily_revenue, 2),
        weekly_growth_pct=round(weekly_growth_pct, 2),
        projections=projections,
    )


def build_narrative(schema: SchemaGuess, metrics: MetricBreakdown, anomalies: List[Anomaly]) -> str:
    parts = [
        f"Detected revenue column: {schema.revenue or 'unknown'}.",
        f"Total revenue {metrics.total_revenue:.2f} with avg daily {metrics.avg_daily_revenue:.2f}.",
        f"Weekly growth {metrics.weekly_growth_pct:.2f}% vs prior week.",
    ]
    if anomalies:
        top = anomalies[0]
        parts.append(f"Anomaly: {top.message}")
    else:
        parts.append("No major anomalies detected this week.")
    # Placeholder where a GPT-5 call would be inserted
    if os.getenv("GPT5_API_KEY"):
        parts.append("GPT-5 hook ready: replace build_narrative with real prompt.")
    return " ".join(parts)


def call_workers_ai(prompt: str) -> Optional[str]:
    account_id = os.getenv("CLOUDFLARE_ACCOUNT_ID")
    api_token = os.getenv("CLOUDFLARE_API_TOKEN")
    model = os.getenv("CLOUDFLARE_MODEL", "@cf/meta/llama-3-8b-instruct")
    if not account_id or not api_token:
        return None
    url = f"https://api.cloudflare.com/client/v4/accounts/{account_id}/ai/run/{model}"
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Content-Type": "application/json",
    }
    res = requests.post(url, headers=headers, json={"prompt": prompt}, timeout=30)
    res.raise_for_status()
    data = res.json()
    return data.get("result", {}).get("response")


def build_pdf(req: PdfRequest) -> bytes:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y, "DataSage Weekly Summary")
    y -= 24

    pdf.setFont("Helvetica", 11)
    pdf.drawString(50, y, f"Total revenue: {req.metrics.total_revenue}")
    y -= 16
    pdf.drawString(50, y, f"Total cost: {req.metrics.total_cost}")
    y -= 16
    pdf.drawString(50, y, f"Net result: {req.metrics.total_profit}")
    y -= 16
    pdf.drawString(50, y, f"Weekly growth: {req.metrics.weekly_growth_pct}%")
    y -= 24

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, "Narrative")
    y -= 16
    pdf.setFont("Helvetica", 10)
    for line in req.narrative.split(". "):
        pdf.drawString(50, y, line.strip())
        y -= 14
        if y < 80:
            pdf.showPage()
            y = height - 50
            pdf.setFont("Helvetica", 10)

    if req.categories:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y, "Top categories")
        y -= 16
        pdf.setFont("Helvetica", 10)
        for cat in req.categories[:6]:
            pdf.drawString(50, y, f"{cat.name}: {cat.total}")
            y -= 14
            if y < 80:
                pdf.showPage()
                y = height - 50
                pdf.setFont("Helvetica", 10)

    if req.anomalies:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y, "Alerts")
        y -= 16
        pdf.setFont("Helvetica", 10)
        for a in req.anomalies[:6]:
            pdf.drawString(50, y, f"{a.date} - {a.message}")
            y -= 14
            if y < 80:
                pdf.showPage()
                y = height - 50
                pdf.setFont("Helvetica", 10)

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.getvalue()


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/upload", response_model=UploadResponse)
async def upload(
    file: UploadFile = File(...),
    date_column: Optional[str] = Form(None),
    revenue_column: Optional[str] = Form(None),
    category_column: Optional[str] = Form(None),
) -> UploadResponse:
    content = await file.read()
    try:
        if file.filename and file.filename.lower().endswith((".xlsx", ".xls")):
            rows = parse_excel(content)
        else:
            rows = parse_csv(content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    schema = apply_overrides(
        guess_schema(rows[0]),
        {
            "date_column": date_column,
            "revenue_column": revenue_column,
            "category_column": category_column,
        },
    )
    columns = list(rows[0].keys())
    if not schema.date or not schema.revenue:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Missing required columns. Please map a date column and an amount/revenue column.",
                "columns": columns,
            },
        )

    return assemble_payload(rows, schema)


@app.post("/explain", response_model=ExplainResponse)
def explain(req: ExplainRequest) -> ExplainResponse:
    prompt = (
        "You are a sharp business analyst. Provide a concise 3-5 sentence explanation. "
        "Include positive, warning, and anomaly cues if present.\n"
        f"Metrics: total_revenue={req.metrics.total_revenue}, total_profit={req.metrics.total_profit}, "
        f"avg_daily={req.metrics.avg_daily_revenue}, weekly_growth={req.metrics.weekly_growth_pct}.\n"
        f"Categories: {[f'{c.name}:{c.total}' for c in req.categories]}.\n"
        f"Anomalies: {[a.message for a in req.anomalies]}.\n"
        f"Period: {req.period or 'latest upload'}."
    )
    tags: List[str] = []
    explanation = call_workers_ai(prompt) or ""
    if not explanation:
        # Fallback heuristic narrative
        growth = req.metrics.weekly_growth_pct
        growth_line = (
            f"Revenue trending up {growth}% week over week." if growth >= 0 else f"Revenue down {abs(growth)}% week over week."
        )
        cat_line = ""
        if req.categories:
            top = max(req.categories, key=lambda c: c.total)
            cat_line = f"Top category: {top.name} at {top.total}."
        anomaly_line = req.anomalies[0].message if req.anomalies else "No anomalies spotted."
        explanation = f"{growth_line} {cat_line} {anomaly_line}"
    if req.metrics.weekly_growth_pct >= 5:
        tags.append("✅ Positive trend")
    if req.metrics.weekly_growth_pct < 0:
        tags.append("⚠️ Warning")
    if req.anomalies:
        tags.append("🔍 Anomaly")
    return ExplainResponse(explanation=explanation, tags=tags or ["Insight ready"])


@app.post("/export/pdf")
def export_pdf(req: PdfRequest) -> Response:
    try:
        pdf_bytes = build_pdf(req)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail="Failed to build PDF") from exc
    headers = {"Content-Disposition": "attachment; filename=summary.pdf"}
    return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)


@app.get("/integrations/google-sheets")
def connect_google_sheets() -> Dict[str, str]:
    return {"status": "ready", "message": "OAuth flow placeholder for Google Sheets."}


def fetch_google_sheet_csv(url: str) -> bytes:
    if not url:
        raise ValueError("Missing Google Sheet URL")
    # Normalize common Google Sheets share links into CSV export links
    if "spreadsheets/d/" in url and "export" not in url:
        # typical share: https://docs.google.com/spreadsheets/d/<id>/edit#gid=0
        try:
            doc_id = url.split("/d/")[1].split("/")[0]
            url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv"
        except Exception:
            # fall through with original url
            ...
    try:
        res = requests.get(url, timeout=15)
        res.raise_for_status()
        return res.content
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Unable to fetch Google Sheet CSV. Ensure the sheet is published to web or use the CSV export link."
        ) from exc


@app.post("/integrations/google-sheets/import", response_model=UploadResponse)
def import_google_sheet(
    sheet_csv_url: str = Form(...),
    date_column: Optional[str] = Form(None),
    revenue_column: Optional[str] = Form(None),
    category_column: Optional[str] = Form(None),
) -> UploadResponse:
    try:
        content = fetch_google_sheet_csv(sheet_csv_url)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    try:
        rows = parse_csv(content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    schema = apply_overrides(
        guess_schema(rows[0]),
        {
            "date_column": date_column,
            "revenue_column": revenue_column,
            "category_column": category_column,
        },
    )
    if not schema.date or not schema.revenue:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Missing required columns. Please map a date column and an amount/revenue column.",
                "columns": list(rows[0].keys()),
            },
        )
    return assemble_payload(rows, schema)


@app.get("/integrations/shopify")
def connect_shopify() -> Dict[str, str]:
    return {"status": "ready", "message": "Shopify OAuth placeholder. Exchange code for access token and store in DB."}


@app.get("/alerts/weekly-summary")
def weekly_summary() -> Dict[str, str]:
    # This would email a PDF; stubbed with simple text.
    return {"status": "ok", "summary": "Weekly PDF summary generator placeholder."}
